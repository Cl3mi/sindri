from openpyxl import Workbook

from app.eval.excel_gold import read_gold_excel, dump_headers


def _sheet(tmp_path, headers, rows, header_row=1, name="g.xlsx"):
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(header_row, col, h)
    for i, row in enumerate(rows, start=header_row + 1):
        for col, v in enumerate(row, start=1):
            ws.cell(i, col, v)
    path = tmp_path / name
    wb.save(path)
    return path


def test_reads_house_style_sheet(tmp_path):
    path = _sheet(tmp_path, ["Pos.", "Merkmal", "Nennmaß", "O-TOL", "U-TOL"],
                  [[1, "Diameter", "20", "0,1", "-0,1"],
                   [2, "Distance", 5.5, None, None]])
    rows = read_gold_excel(path)
    assert rows[1] == {"char_type": "Diameter", "nominal": "20",
                       "upper_tol": "0,1", "lower_tol": "-0,1", "raw": ""}
    assert rows[2]["nominal"] == "5.5"       # numeric cell -> canonical string
    assert rows[2]["upper_tol"] == ""        # None -> empty


def test_header_aliases_and_offset_header_row(tmp_path):
    path = _sheet(tmp_path, ["Position", "Characteristic", "Nominal value",
                             "Upper-tol", "Lower-tol"],
                  [[7, "Radius", "2", "0", ""]], header_row=3)
    rows = read_gold_excel(path)
    assert rows[7]["char_type"] == "Radius"


def test_missing_pos_column_raises(tmp_path):
    path = _sheet(tmp_path, ["Foo", "Bar"], [[1, 2]])
    try:
        read_gold_excel(path)
        assert False, "expected ValueError"
    except ValueError as e:
        assert "Pos" in str(e)


def test_dump_headers_reports_detected_row(tmp_path):
    path = _sheet(tmp_path, ["Pos.", "Merkmal", "Nennmaß", "O-TOL", "U-TOL"],
                  [[1, "Diameter", "20", "", ""]])
    info = dump_headers(path)
    assert info["header_row"] == 1
    assert "Merkmal" in info["headers"]
    assert info["n_rows"] == 1


def test_dump_headers_scans_every_sheet_and_reports_where_pos_lives(tmp_path):
    """Real workbooks carry a cover sheet before the measurement table, and the
    header sits well below row 1. The inspector must say which sheet and which
    row, otherwise a failure is undiagnosable."""
    from openpyxl import Workbook
    wb = Workbook()
    cover = wb.active
    cover.title = "Deckblatt"
    cover.cell(1, 1, "irrelevant")
    ws = wb.create_sheet("Messprotokoll")
    ws.cell(7, 1, "Nr.")
    ws.cell(7, 2, "Sollwert")
    ws.cell(8, 1, 1)
    ws.cell(8, 2, "20")
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    info = dump_headers(path)
    assert info["n_sheets"] == 2
    assert info["sheet_names"] == ["Deckblatt", "Messprotokoll"]
    found = {s["sheet"]: s["pos_row"] for s in info["scan"]}
    assert found["Deckblatt"] is None
    assert found["Messprotokoll"] == 7


def test_bilingual_single_cell_captions_are_matched(tmp_path):
    """The real sheets carry German AND English in ONE header cell, because the
    two header rows of app/excel.py get merged. Exact-match on 'pos.' misses
    'Pos.\\nPos.', which is why 97/100 client sheets failed to parse."""
    path = _sheet(tmp_path,
                  ["Pos.\nPos.", "Merkmal\nCharacteristic",
                   "Nennmaß\nNominal value", "O-TOL\nUpper-tol",
                   "U-TOL\nLower-tol"],
                  [[1, "Diameter", "20", "0,1", "-0,1"]])
    rows = read_gold_excel(path)
    assert rows[1] == {"char_type": "Diameter", "nominal": "20",
                       "upper_tol": "0,1", "lower_tol": "-0,1", "raw": ""}


def test_header_below_the_shallow_scan_is_found(tmp_path):
    """One sheet family puts the table header at row 17, under a tall
    title/metadata block. A 12-row scan silently rejects the whole workbook."""
    path = _sheet(tmp_path,
                  ["Pos.\nPos.", "Merkmal\nCharacteristic",
                   "Nennmaß\nNominal value", "O-TOL\nUpper-tol",
                   "U-TOL\nLower-tol"],
                  [[1, "Diameter", "20", "", ""]], header_row=17)
    rows = read_gold_excel(path)
    assert rows[1]["char_type"] == "Diameter"


def test_german_measurement_report_captions_are_matched(tmp_path):
    """The second family of sheets uses Sollwert/Obere/Untere instead."""
    path = _sheet(tmp_path, ["Nr.", "Maß", "Sollwert", "Obere", "Untere"],
                  [[3, "Abstand", "12,5", "0,2", "-0,2"]], header_row=7)
    rows = read_gold_excel(path)
    assert rows[3]["nominal"] == "12,5"
    assert rows[3]["upper_tol"] == "0,2"
    assert rows[3]["lower_tol"] == "-0,2"


def test_sheet_vocabulary_collects_caption_like_strings_only(tmp_path):
    """Cross-document frequency is what makes this safe: a caption repeats in
    many workbooks, a measured value does not. Numbers are dropped outright."""
    from openpyxl import Workbook
    from app.eval.excel_gold import sheet_vocabulary
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Messbericht Nr.")
    ws.cell(2, 1, "Sollwert")
    ws.cell(2, 2, 20.5)              # numeric value -> excluded
    ws.cell(3, 1, "12,5")            # numeric-as-text -> excluded
    ws.cell(3, 2, "x")               # too short -> excluded
    ws.cell(99, 1, "Deep Caption")   # below the scan window -> excluded
    path = tmp_path / "v.xlsx"
    wb.save(path)

    vocab = sheet_vocabulary(path, max_rows=25)
    assert "Messbericht Nr." in vocab
    assert "Sollwert" in vocab
    assert "20.5" not in vocab and "12,5" not in vocab
    assert "x" not in vocab
    assert "Deep Caption" not in vocab


def test_dump_headers_reports_duplicate_pos(tmp_path):
    path = _sheet(tmp_path, ["Pos.", "Merkmal", "Nennmaß", "O-TOL", "U-TOL"],
                  [[1, "Diameter", "20", "", ""],
                   [1, "Distance", "7", "", ""],
                   [2, "Radius", "2", "", ""]])
    info = dump_headers(path)
    assert info["duplicate_pos"] == [1]
    assert info["n_rows"] == 2
