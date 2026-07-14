import fitz
from openpyxl import load_workbook

from app.eval.models import GoldCharacteristic
from app.eval.synthetic import make_synthetic_doc
from app.eval.balloons import recover_balloons

RECORDS = [
    GoldCharacteristic(balloon=1, position_pt=(120.0, 90.0),
                       char_type="Diameter", nominal="20",
                       upper_tol="0,1", lower_tol="-0,1"),
    GoldCharacteristic(balloon=2, position_pt=(340.0, 200.0),
                       char_type="Distance", nominal="5,5"),
    GoldCharacteristic(balloon=3, position_pt=(500.0, 320.0),
                       char_type="Radius", nominal="2", upper_tol="0"),
]


def test_synthetic_doc_produces_recoverable_balloons_and_gold_excel(tmp_path):
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")
    assert pdf.name == "SYN1.pdf" and xlsx.name == "SYN1.xlsx"

    balloons = recover_balloons(pdf)
    assert sorted(b.number for b in balloons) == [1, 2, 3]

    wb = load_workbook(xlsx)
    ws = wb.active
    header = [ws.cell(1, c).value for c in range(1, 6)]
    assert header == ["Pos.", "Merkmal", "Nennmaß", "O-TOL", "U-TOL"]
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 3).value == "20"
    assert ws.cell(3, 3).value == "5,5"


def test_synthetic_page_size_defaults_to_a3_landscape(tmp_path):
    pdf, _ = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN2")
    doc = fitz.open(pdf)
    assert round(doc[0].rect.width) == 1191 and round(doc[0].rect.height) == 842
    doc.close()
