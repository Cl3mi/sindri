import fitz

from app.eval.models import GoldCharacteristic
from app.eval.synthetic import make_synthetic_doc
from app.eval.ingest import build_gold_doc

RECORDS = [
    GoldCharacteristic(balloon=1, position_pt=(120.0, 90.0),
                       char_type="Diameter", nominal="20",
                       upper_tol="0,1", lower_tol="-0,1"),
    GoldCharacteristic(balloon=2, position_pt=(340.0, 200.0),
                       char_type="Distance", nominal="5,5"),
]


def test_join_recovers_positions_and_values(tmp_path):
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")
    gold = build_gold_doc(pdf, xlsx, doc_id="SYN1")
    assert gold.doc_id == "SYN1"
    assert round(gold.page_rect[2]) == 1191
    by_num = {c.balloon: c for c in gold.characteristics}
    assert set(by_num) == {1, 2}
    assert by_num[1].nominal == "20" and by_num[1].char_type == "Diameter"
    x, y = by_num[2].position_pt
    assert abs(x - 340.0) < 3 and abs(y - 200.0) < 3
    assert gold.provenance["join_rate"] == 1.0


def _resized_copy(src, dst, factor):
    """The same drawing on a sheet `factor` times the size, content scaled to
    fit -- which is the relationship measured between the delivered stamped
    exports and the clean originals."""
    doc = fitz.open(src)
    rect = doc[0].rect
    out = fitz.open()
    page = out.new_page(width=rect.width * factor, height=rect.height * factor)
    page.show_pdf_page(page.rect, doc, 0)
    doc.close()
    out.save(dst)
    out.close()
    return dst


def test_gold_positions_land_in_the_originals_page_space(tmp_path):
    """Balloons only exist on the STAMPED drawing, but the pipeline reads the
    CLEAN original -- and on 14 of 20 dev documents those two sheets have
    different extents, which put every gold position in a coordinate space the
    predictions never occupy. Recovery still happens on the stamped sheet;
    positions and page_rect are reported in the original's space."""
    stamped, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")
    original = _resized_copy(stamped, tmp_path / "SYN1_orig.pdf", 2.0)

    gold = build_gold_doc(stamped, xlsx, doc_id="SYN1", target_pdf=original)

    # page_rect is the ORIGINAL's, so scoring compares like with like
    assert round(gold.page_rect[2]) == round(1191 * 2)
    # and every balloon moved with it: (340, 200) on the stamped sheet is
    # (680, 400) on a sheet twice the size
    by_num = {c.balloon: c for c in gold.characteristics}
    x, y = by_num[2].position_pt
    assert abs(x - 680.0) < 6 and abs(y - 400.0) < 6
    # traceable, but in provenance so it stays out of gold_hash
    assert gold.provenance["target_scale"] == [2.0, 2.0]


def test_a_document_with_no_original_drops_its_positions(tmp_path):
    """One drawing in the corpus (976d3c0d) has no clean original. Keeping its
    balloons in the stamped sheet's coordinate space silently reintroduces the
    exact fault --originals exists to remove, and charges the pipeline for
    "missing" callouts at coordinates it was never shown. A row with no position
    is matched on value instead and counts as unlocated -- which is the truth:
    no detection change can ever fix it."""
    stamped, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")

    gold = build_gold_doc(stamped, xlsx, doc_id="SYN1", target_pdf=None,
                          require_target=True)

    assert all(c.position_pt is None for c in gold.characteristics)
    assert gold.provenance["target_missing"] is True
    # and it is still real gold -- the rows are not dropped
    assert {c.balloon for c in gold.characteristics} == {1, 2}


def test_gold_is_unchanged_when_no_originals_are_given(tmp_path):
    """Default off: omitting target_pdf must reproduce byte-identical gold, so
    adding this cannot silently move an existing corpus."""
    stamped, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")

    plain = build_gold_doc(stamped, xlsx, doc_id="SYN1")
    explicit_none = build_gold_doc(stamped, xlsx, doc_id="SYN1", target_pdf=None)

    assert plain.gold_hash() == explicit_none.gold_hash()
    assert round(plain.page_rect[2]) == 1191


def test_same_sized_originals_leave_gold_hash_untouched(tmp_path):
    """The 6 documents whose sheets already agree must not move at all -- a
    same-size target is an identity transform, not a no-op-shaped rewrite."""
    stamped, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN1")
    same = _resized_copy(stamped, tmp_path / "SYN1_same.pdf", 1.0)

    plain = build_gold_doc(stamped, xlsx, doc_id="SYN1")
    targeted = build_gold_doc(stamped, xlsx, doc_id="SYN1", target_pdf=same)

    assert plain.gold_hash() == targeted.gold_hash()


def _flatten(src, dst, dpi=200):
    """Raster-only copy: no text layer survives, matching the delivered
    drawings whose balloon numbers defeat text recovery."""
    import fitz
    doc = fitz.open(src)
    rect = doc[0].rect
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72))
    doc.close()
    out = fitz.open()
    page = out.new_page(width=rect.width, height=rect.height)
    page.insert_image(page.rect, pixmap=pix)
    out.save(dst)
    out.close()
    return dst


def test_cv_fallback_fills_positions_the_text_layer_cannot(tmp_path):
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYNCV")
    flat = _flatten(pdf, tmp_path / "flat.pdf")

    plain = build_gold_doc(flat, xlsx, doc_id="SYNCV")
    assert all(c.position_pt is None for c in plain.characteristics)
    assert plain.provenance["without_position"] == len(RECORDS)

    with_cv = build_gold_doc(flat, xlsx, doc_id="SYNCV", use_cv=True)
    located = [c for c in with_cv.characteristics if c.position_pt is not None]
    assert located, "CV fallback recovered no positions"
    assert with_cv.provenance["recovered_by_cv"] == len(located)
    # every row is still gold either way
    assert len(with_cv.characteristics) == len(RECORDS)


def test_provenance_reports_the_kind_of_rows_that_lack_a_balloon(tmp_path):
    """Decides whether unlocated rows are a detection failure or a data
    reality: a FAI sheet lists material and general notes that were never
    ballooned. char_type is a shared category label, not a measurement."""
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN4")
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb.active
    ws.cell(4, 1, 9); ws.cell(4, 2, "Material"); ws.cell(4, 3, "S235")
    wb.save(xlsx)

    gold = build_gold_doc(pdf, xlsx, doc_id="SYN4")
    assert gold.provenance["unlocated_char_types"] == {"Material": 1}


def test_excel_rows_without_a_balloon_stay_in_the_gold(tmp_path):
    """A characteristic whose balloon could not be located is still a
    characteristic. Dropping it would make the eval blind to ever missing it —
    on the real corpus that silently discarded 17% of the ground truth."""
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN3")
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb.active
    ws.cell(4, 1, 9); ws.cell(4, 2, "Distance"); ws.cell(4, 3, "7")
    wb.save(xlsx)

    gold = build_gold_doc(pdf, xlsx, doc_id="SYN3")
    by_num = {c.balloon: c for c in gold.characteristics}
    assert set(by_num) == {1, 2, 9}              # row 9 retained
    assert by_num[9].position_pt is None         # but with no position
    assert by_num[1].position_pt is not None
    assert gold.provenance["without_position"] == 1


def test_unjoined_rows_and_balloons_recorded_not_dropped_silently(tmp_path):
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN2")
    # Excel has a row 9 with no balloon on the page
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb.active
    ws.cell(4, 1, 9); ws.cell(4, 2, "Distance"); ws.cell(4, 3, "7")
    wb.save(xlsx)
    gold = build_gold_doc(pdf, xlsx, doc_id="SYN2")
    assert gold.provenance["excel_only"] == [9]
    assert gold.provenance["pdf_only"] == []
    assert gold.provenance["join_rate"] < 1.0
    # The unlocated row is kept as gold; only its POSITION is missing.
    assert {c.balloon for c in gold.characteristics} == {1, 2, 9}
    assert {c.balloon for c in gold.characteristics
            if c.position_pt is not None} == {1, 2}


def test_duplicate_balloon_numbers_surfaced_in_provenance(tmp_path):
    pdf, xlsx = make_synthetic_doc(RECORDS, tmp_path, doc_id="SYN3")
    # drawings repeat balloon "1" at a second position (e.g. a second view)
    doc = fitz.open(pdf)
    page = doc[0]
    x, y = 700.0, 600.0
    page.draw_circle(fitz.Point(x, y), 9.0, color=(0, 0, 1), width=1.5)
    page.insert_text(fitz.Point(x - 5, y + 4), "1", fontsize=10, color=(0, 0, 1))
    doc.saveIncr()
    doc.close()

    gold = build_gold_doc(pdf, xlsx, doc_id="SYN3")
    assert gold.provenance["duplicate_balloons"] == [1]
