from openpyxl import load_workbook
from app.models import Characteristic, TitleField
from app.excel import write_workbook

def test_write_workbook(tmp_path):
    rows = [
        Characteristic(pos=2, char_type="Distance", nominal="3,2", upper_tol="0,05", lower_tol="-0,05"),
        Characteristic(pos=1, char_type="Distance", nominal="1,2", upper_tol="0,1", lower_tol="-0,1"),
    ]
    out = tmp_path / "out.xlsx"
    write_workbook(rows, out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(1, 1).value == "Pos."
    assert ws.cell(1, 2).value == "Merkmal"
    assert ws.cell(2, 2).value == "Characteristic"
    assert ws.cell(3, 1).value == 1
    assert ws.cell(3, 3).value == "1,2"
    assert ws.cell(4, 1).value == 2


def test_notes_sheet_created_when_note_block_passed(tmp_path):
    from openpyxl import load_workbook
    from app.excel import write_workbook
    from app.models import Note, NoteBlock, Characteristic

    nb = NoteBlock(region=(0, 0, 10, 10), notes=[
        Note(pos=101, text_en="CONTACT AREA NOTES", text_de="KONTAKTBEREICH HINWEISE"),
        Note(pos=1, parent_pos=101, sub_index=1, text_en="PLANARITY", text_de="EBENHEIT"),
        Note(pos=2, parent_pos=101, sub_index=2, text_en="SURFACE", text_de="OBERFLAECHE"),
        Note(pos=102, text_en="PART FREE OF GREASE", text_de="OHNE FETT"),
    ])
    out = tmp_path / "x.xlsx"
    write_workbook([Characteristic(pos=1, char_type="Distance", nominal="1,2")],
                   out, notes=nb)
    wb = load_workbook(out)
    assert "Notes" in wb.sheetnames
    ws = wb["Notes"]
    # Headers
    assert ws.cell(1, 1).value == "Pos"
    assert ws.cell(1, 2).value == "English"
    assert ws.cell(1, 3).value == "German"
    # Rows in order, with sub-bullet pos formatted as "101.1"
    assert ws.cell(2, 1).value == "101"
    assert ws.cell(3, 1).value == "101.1"
    assert ws.cell(4, 1).value == "101.2"
    assert ws.cell(5, 1).value == "102"


def test_notes_sheet_absent_when_no_notes_passed(tmp_path):
    from openpyxl import load_workbook
    from app.excel import write_workbook
    from app.models import Characteristic

    out = tmp_path / "x.xlsx"
    write_workbook([Characteristic(pos=1, char_type="Distance", nominal="1,2")],
                   out)
    wb = load_workbook(out)
    assert "Notes" not in wb.sheetnames


def test_workbook_has_title_block_sheet(tmp_path):
    from app.excel import write_workbook
    fields = [
        TitleField(label="Sheet / Blatt", label_en="Sheet", label_de="Blatt",
                   value="1/1"),
        TitleField(label="Scale / Maßstab", label_en="Scale", label_de="Maßstab",
                   value="5:1"),
    ]
    out = tmp_path / "wb.xlsx"
    write_workbook([], out, title_block=fields)
    wb = load_workbook(out)
    assert "Title Block" in wb.sheetnames
    ws = wb["Title Block"]
    assert [c.value for c in ws[1]] == ["Label (EN)", "Label (DE)", "Value"]
    assert ws.cell(2, 1).value == "Sheet" and ws.cell(2, 3).value == "1/1"
    assert ws.cell(3, 2).value == "Maßstab"


def test_workbook_omits_title_block_sheet_when_empty(tmp_path):
    from app.excel import write_workbook
    out = tmp_path / "wb.xlsx"
    write_workbook([], out, title_block=[])
    wb = load_workbook(out)
    assert "Title Block" not in wb.sheetnames


def test_write_workbook_creates_marks_sheet(tmp_path):
    from openpyxl import load_workbook
    from app.models import Characteristic, Mark, MarkBlock
    from app.excel import write_workbook

    rows = [Characteristic(pos=1, char_type="Distance", nominal="10")]
    marks = MarkBlock(region=(0, 0, 100, 100), marks=[
        Mark(pos=101, text_en="EN-A", text_de="DE-A"),
        Mark(pos=102, text_en="EN-B", text_de="DE-B"),
    ])
    out = tmp_path / "out.xlsx"
    write_workbook(rows, out, marks=marks)

    wb = load_workbook(out)
    assert "Marks" in wb.sheetnames
    ws = wb["Marks"]
    # row 1 = headers; row 2+ = marks
    assert ws.cell(1, 1).value == "Pos"
    assert ws.cell(1, 2).value == "English"
    assert ws.cell(1, 3).value == "German"
    assert ws.cell(2, 1).value == "101"
    assert ws.cell(2, 2).value == "EN-A"
    assert ws.cell(2, 3).value == "DE-A"
    assert ws.cell(3, 1).value == "102"


def test_write_workbook_omits_marks_sheet_when_marks_none(tmp_path):
    from openpyxl import load_workbook
    from app.models import Characteristic
    from app.excel import write_workbook

    out = tmp_path / "out.xlsx"
    write_workbook([Characteristic(pos=1)], out)
    wb = load_workbook(out)
    assert "Marks" not in wb.sheetnames


def test_marks_sheet_ordered_before_notes_when_both_present(tmp_path):
    from openpyxl import load_workbook
    from app.models import Characteristic, Mark, MarkBlock, Note, NoteBlock
    from app.excel import write_workbook

    out = tmp_path / "out.xlsx"
    write_workbook(
        [Characteristic(pos=1)], out,
        notes=NoteBlock(region=(0, 0, 1, 1), notes=[Note(pos=101)]),
        marks=MarkBlock(region=(0, 0, 1, 1), marks=[Mark(pos=101)]),
    )
    wb = load_workbook(out)
    assert wb.sheetnames == ["Inspection", "Marks", "Notes"]
