"""Build a synthetic 'client deliverable' (ballooned PDF + gold Excel) from
known GoldCharacteristic records. Test bed for the whole gold pipeline: the
balloons match the client's expected vector encoding (circle + centered number)
and the Excel mirrors the expected schema (same header vocabulary as
app/excel.py — the client sheets came out of the same inspection workflow;
Task 13 confirms against real files and only excel_gold's schema config needs
to change if they differ)."""
from pathlib import Path
from typing import List, Tuple

import fitz
from openpyxl import Workbook

from app.eval.models import GoldCharacteristic

_A3_LANDSCAPE = (1191, 842)   # points
_HEADERS = ["Pos.", "Merkmal", "Nennmaß", "O-TOL", "U-TOL"]


def make_synthetic_doc(records: List[GoldCharacteristic], out_dir,
                       doc_id: str = "SYN",
                       page_size: Tuple[int, int] = _A3_LANDSCAPE,
                       ) -> Tuple[Path, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / f"{doc_id}.pdf"
    xlsx_path = out_dir / f"{doc_id}.xlsx"

    doc = fitz.open()
    page = doc.new_page(width=page_size[0], height=page_size[1])
    for r in records:
        x, y = r.position_pt
        page.draw_circle(fitz.Point(x, y), 9.0, color=(0, 0, 1), width=1.5)
        page.insert_text(fitz.Point(x - 5, y + 4), str(r.balloon),
                         fontsize=10, color=(0, 0, 1))
        # the callout text the balloon points at, offset like a real drawing
        label = f"{r.nominal} {r.upper_tol} {r.lower_tol}".strip()
        page.insert_text(fitz.Point(x + 14, y + 4), label, fontsize=8)
    doc.save(pdf_path)
    doc.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection"
    for col, h in enumerate(_HEADERS, start=1):
        ws.cell(1, col, h)
    for i, r in enumerate(sorted(records, key=lambda r: r.balloon), start=2):
        ws.cell(i, 1, r.balloon)
        ws.cell(i, 2, r.char_type)
        ws.cell(i, 3, r.nominal)
        ws.cell(i, 4, r.upper_tol)
        ws.cell(i, 5, r.lower_tol)
    wb.save(xlsx_path)
    return pdf_path, xlsx_path
