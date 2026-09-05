"""Read a client gold Excel into {balloon_number: field dict}.

Schema is an adapter around COLUMN_ALIASES + header auto-detection: the header
row is found by scanning the first 12 rows for a 'Pos' alias. When the real
corpus arrives (Task 13), run `dump_headers` over all files; if the layout
differs, extend COLUMN_ALIASES — nothing else changes.

Numeric cells are canonicalized through normalize.canon_value at read time so
'5,5' (text) and 5.5 (float cell) ingest identically.
"""
from typing import Dict, List, Optional

from openpyxl import load_workbook

from app.eval.normalize import canon_value

# canonical field -> header aliases (matched casefolded/stripped)
COLUMN_ALIASES: Dict[str, List[str]] = {
    "pos": ["pos.", "pos", "position", "nr.", "nr", "ballon", "balloon"],
    # Aliases confirmed against the real corpus (2026-08-17): the bilingual
    # house sheets use Merkmal/Nennmaß/O-TOL/U-TOL, the measurement-report
    # family uses Maß/Sollwert/Obere/Untere.
    "char_type": ["merkmal", "characteristic", "typ", "type", "maß", "mass"],
    "nominal": ["nennmaß", "nennmass", "nominal value", "nominal", "soll",
                "sollwert"],
    "upper_tol": ["o-tol", "upper-tol", "oberes abmaß", "upper tol", "otol",
                  "obere", "oberes"],
    "lower_tol": ["u-tol", "lower-tol", "unteres abmaß", "lower tol", "utol",
                  "untere", "unteres"],
    "raw": ["raw", "text", "bemerkung", "remark"],
}
# Confirmed against the real corpus (2026-08-17): 70 sheets put the header at
# row 10 and 27 at row 17, under a tall title/metadata block. A 12-row scan
# silently rejected a quarter of the corpus.
_MAX_HEADER_SCAN = 25


def _norm_header(v) -> str:
    return " ".join(str(v or "").split()).casefold()


def _header_keys(v) -> set:
    """Every way one header cell might be looked up.

    Client sheets carry BOTH languages in a single cell ("Pos.\\nPos.",
    "Nennmaß\\nNominal value") because the two header rows written by
    app/excel.py end up merged. Matching only the flattened text misses them,
    so each line is a candidate key too."""
    raw = str(v or "")
    keys = {_norm_header(raw)}
    for line in raw.splitlines():
        line = _norm_header(line)
        if line:
            keys.add(line)
    keys.discard("")
    return keys


def _find_header(ws, max_scan: int = _MAX_HEADER_SCAN):
    """Return (header_row, {field: column}) or raise ValueError."""
    pos_aliases = set(COLUMN_ALIASES["pos"])
    for row in range(1, min(max_scan, ws.max_row) + 1):
        headers = {}
        for c in range(1, ws.max_column + 1):
            for key in _header_keys(ws.cell(row, c).value):
                headers.setdefault(key, c)
        if not (pos_aliases & set(headers)):
            continue
        cols = {}
        for field, aliases in COLUMN_ALIASES.items():
            for a in aliases:
                if a in headers:
                    cols[field] = headers[a]
                    break
        return row, cols
    raise ValueError(f"no header row with a 'Pos' column found in {ws.title!r} "
                     f"(scanned {_MAX_HEADER_SCAN} rows)")


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return canon_value(v)
    return str(v).strip()


def read_gold_excel(path, sheet: Optional[str] = None) -> Dict[int, dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    header_row, cols = _find_header(ws)
    out: Dict[int, dict] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        pos_v = ws.cell(row, cols["pos"]).value
        if pos_v is None or not str(pos_v).strip():
            continue
        try:
            balloon = int(float(str(pos_v).replace(",", ".")))
        except ValueError:
            continue                      # sub-header / footer rows
        out[balloon] = {
            field: _cell_str(ws.cell(row, col).value)
            for field, col in cols.items() if field != "pos"
        }
        for field in ("char_type", "nominal", "upper_tol", "lower_tol", "raw"):
            out[balloon].setdefault(field, "")
    return out


_DEEP_SCAN = 40


def _scan_all_sheets(wb) -> list:
    """Diagnostic: for every worksheet, the row where a 'Pos'-alias header sits
    (searching deeper than the reader does), or None. Distinguishes 'wrong
    sheet' from 'header too deep' from 'no such column at all'."""
    out = []
    for ws in wb.worksheets:
        row = None
        try:
            row, _ = _find_header(ws, max_scan=_DEEP_SCAN)
        except ValueError:
            pass
        out.append({"sheet": ws.title, "pos_row": row, "max_row": ws.max_row})
    return out


def sheet_vocabulary(path, max_rows: int = 25) -> set:
    """Caption-like strings from the top of every sheet.

    Used only in aggregate: the runner reports strings that appear in MANY
    workbooks, which are by construction shared captions rather than per-part
    measurements. Anything numeric, very short, or letter-free is dropped, so a
    measured value cannot reach the summary even if it repeated."""
    wb = load_workbook(path, data_only=True)
    vocab = set()
    for ws in wb.worksheets:
        for row in range(1, min(max_rows, ws.max_row) + 1):
            for col in range(1, min(ws.max_column, 40) + 1):
                value = ws.cell(row, col).value
                if not isinstance(value, str):
                    continue
                text = " ".join(value.split())
                if len(text) < 2 or len(text) > 40:
                    continue
                if not any(ch.isalpha() for ch in text):
                    continue
                if _try_number(text) is not None:
                    continue
                vocab.add(text)
    return vocab


def _try_number(text: str):
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def dump_headers(path) -> dict:
    """Day-one inspection: which sheet, header row and labels does this use?"""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    scan = _scan_all_sheets(wb)
    extra = {"n_sheets": len(wb.worksheets),
             "sheet_names": [w.title for w in wb.worksheets],
             "scan": scan}
    try:
        header_row, cols = _find_header(ws)
        rows = read_gold_excel(path)
        counts: Dict[int, int] = {}
        for row in range(header_row + 1, ws.max_row + 1):
            pos_v = ws.cell(row, cols["pos"]).value
            if pos_v is None or not str(pos_v).strip():
                continue
            try:
                balloon = int(float(str(pos_v).replace(",", ".")))
            except ValueError:
                continue
            counts[balloon] = counts.get(balloon, 0) + 1
        return {
            "file": str(path),
            "sheet": ws.title,
            "header_row": header_row,
            "headers": [str(ws.cell(header_row, c).value)
                        for c in range(1, ws.max_column + 1)
                        if ws.cell(header_row, c).value is not None],
            "mapped_fields": sorted(cols),
            "n_rows": len(rows),
            "duplicate_pos": sorted(n for n, c in counts.items() if c > 1),
            **extra,
        }
    except ValueError as e:
        return {"file": str(path), "sheet": ws.title, "error": str(e), **extra}
