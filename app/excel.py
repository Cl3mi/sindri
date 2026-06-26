from pathlib import Path
from typing import Iterable, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from app.models import Characteristic, NoteBlock, TitleField

HEADERS = [
    ("Pos.", "Pos."),
    ("Merkmal", "Characteristic"),
    ("Nennmaß", "Nominal value"),
    ("O-TOL", "Upper-tol"),
    ("U-TOL", "Lower-tol"),
]

_thin = Side(style="thin")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_center = Alignment(horizontal="center", vertical="center")


def _write_characteristics_sheet(ws, rows: Iterable[Characteristic]) -> None:
    for col, (de, en) in enumerate(HEADERS, start=1):
        top = ws.cell(1, col, de)
        bot = ws.cell(2, col, en)
        for cell in (top, bot):
            cell.font = Font(bold=True)
            cell.alignment = _center
            cell.border = _border

    ordered = sorted(rows, key=lambda c: c.pos)
    for i, c in enumerate(ordered, start=3):
        values = [c.pos, c.char_type, c.nominal, c.upper_tol, c.lower_tol]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(i, col, v)
            cell.alignment = _center
            cell.border = _border

    widths = [8, 18, 16, 12, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w


def _write_notes_sheet(ws, block: NoteBlock) -> None:
    headers = ["Pos", "English", "German"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.alignment = _center
        cell.border = _border
    # Render notes in source order. Sub-bullets show "<parent>.<sub>" as Pos.
    for i, n in enumerate(block.notes, start=2):
        if n.parent_pos is not None and n.sub_index is not None:
            pos_label = f"{n.parent_pos}.{n.sub_index}"
        else:
            pos_label = f"{n.pos}"
        ws.cell(i, 1, pos_label)
        ws.cell(i, 2, n.text_en)
        ws.cell(i, 3, n.text_de)
    for col, w in enumerate([10, 48, 48], start=1):
        ws.column_dimensions[chr(64 + col)].width = w


def _write_title_block_sheet(ws, fields) -> None:
    headers = ["Label (EN)", "Label (DE)", "Value"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.alignment = _center
        cell.border = _border
    for i, f in enumerate(fields, start=2):
        ws.cell(i, 1, f.label_en)
        ws.cell(i, 2, f.label_de)
        ws.cell(i, 3, f.value)
    for col, w in enumerate([22, 22, 40], start=1):
        ws.column_dimensions[chr(64 + col)].width = w


def write_workbook(rows: Iterable[Characteristic], path: Path,
                   notes: Optional[NoteBlock] = None,
                   title_block: Optional[Iterable[TitleField]] = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection"
    _write_characteristics_sheet(ws, rows)
    if notes is not None and notes.notes:
        _write_notes_sheet(wb.create_sheet("Notes"), notes)
    if title_block:
        _write_title_block_sheet(wb.create_sheet("Title Block"), list(title_block))
    path = Path(path)
    wb.save(path)
